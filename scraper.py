#!/usr/bin/env python3
"""
scraper.py — Scraper de notícias para fontes sem RSS público
Matinal · Monitor de Notícias
"""

import re
import time
import datetime
import logging
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Optional
from urllib.parse import urljoin, quote

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger(__name__)

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/122.0.0.0 Safari/537.36'
    ),
    'Accept-Language': 'pt-BR,pt;q=0.9,en;q=0.8',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
}
TIMEOUT = 20
DELAY   = 0.8   # segundos entre requests


# ─────────────────────────────────────────────────────────────────────────────
# Estrutura de dados
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Article:
    source: str
    title: str
    link: str
    date: Optional[datetime.datetime] = None
    description: str = ''
    group: str = 'Sem RSS'


# ─────────────────────────────────────────────────────────────────────────────
# Utilitários
# ─────────────────────────────────────────────────────────────────────────────

def get_soup(url: str, **kw) -> Optional[BeautifulSoup]:
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, **kw)
        r.raise_for_status()
        return BeautifulSoup(r.text, 'html.parser')
    except requests.exceptions.HTTPError as e:
        log.warning(f"HTTP {e.response.status_code} em {url}")
    except requests.exceptions.ConnectionError:
        log.warning(f"Erro de conexão em {url}")
    except requests.exceptions.Timeout:
        log.warning(f"Timeout em {url}")
    except Exception as e:
        log.warning(f"Erro em {url}: {e}")
    return None


def abs_url(href: str, base: str) -> str:
    if not href:
        return ''
    href = href.strip()
    if href.startswith('http'):
        return href
    if href.startswith('//'):
        return 'https:' + href
    return urljoin(base, href)


def parse_rfc2822(s: str) -> Optional[datetime.datetime]:
    try:
        from email.utils import parsedate_to_datetime
        return parsedate_to_datetime(s)
    except Exception:
        return None


def dedupe(articles: list[Article]) -> list[Article]:
    seen: set[str] = set()
    out: list[Article] = []
    for a in articles:
        key = a.title.lower().strip()[:70]
        if key and key not in seen:
            seen.add(key)
            out.append(a)
    return out


def extract_titles(soup: BeautifulSoup, selectors: list[str],
                   base_url: str, source: str, group: str,
                   min_len: int = 20) -> list[Article]:
    """Extrai títulos/links usando lista de seletores CSS (em ordem de prioridade)."""
    articles: list[Article] = []
    seen_links: set[str] = set()
    for sel in selectors:
        for a in soup.select(sel):
            title = a.get_text(separator=' ', strip=True)
            link  = abs_url(a.get('href', ''), base_url)
            if not title or len(title) < min_len:
                continue
            if link in seen_links:
                continue
            seen_links.add(link)
            articles.append(Article(source=source, title=title,
                                    link=link, group=group))
    return articles


# ─────────────────────────────────────────────────────────────────────────────
# Scrapers individuais
# ─────────────────────────────────────────────────────────────────────────────

def scrape_nonada() -> list[Article]:
    """nonada.com.br — jornalismo independente, WordPress."""
    BASE = 'https://nonada.com.br'
    soup = get_soup(BASE)
    if not soup:
        return []
    articles = extract_titles(soup, [
        'article h2 a', 'article h3 a',
        '.entry-title a', '.post-title a',
        'h2.wp-block-post-title a', 'h3.wp-block-post-title a',
        '.is-style-headline a',
    ], BASE, 'Nonada', 'RS')
    # fallback: qualquer link de artigo na home
    if not articles:
        for a in soup.find_all('a', href=True):
            link = abs_url(a['href'], BASE)
            if re.match(r'https://nonada\.com\.br/\d{4}/\d{2}/', link):
                title = a.get_text(strip=True)
                if len(title) >= 20:
                    articles.append(Article('Nonada', title, link, group='RS'))
    return dedupe(articles)


def scrape_sul21() -> list[Article]:
    """sul21.com.br — jornalismo independente, WordPress."""
    BASE = 'https://sul21.com.br'
    soup = get_soup(BASE)
    if not soup:
        return []
    return dedupe(extract_titles(soup, [
        'article h2 a', 'article h3 a',
        '.entry-title a', '.post-title a',
        'h2 a[href*="sul21.com.br"]', 'h3 a[href*="sul21.com.br"]',
    ], BASE, 'Sul21', 'RS'))


def scrape_ufrgs_jornal() -> list[Article]:
    """ufrgs.br/jornal — Jornal da Universidade (UFRGS)."""
    BASE = 'https://www.ufrgs.br'
    urls = ['https://www.ufrgs.br/jornal/', 'https://www.ufrgs.br/jornal/category/noticias/']
    articles: list[Article] = []
    for url in urls:
        soup = get_soup(url)
        if not soup:
            continue
        articles += extract_titles(soup, [
            'article h2 a', 'article h3 a',
            '.entry-title a', '.post-title a',
            'h2.title a', 'h3.title a',
            '.views-field-title a',
        ], BASE, 'UFRGS Jornal', 'RS')
        time.sleep(DELAY)
    return dedupe(articles)


def scrape_ufrgs_noticias() -> list[Article]:
    """ufrgs.br/site/noticias — Notícias institucionais da UFRGS."""
    BASE = 'https://www.ufrgs.br'
    urls = [
        'https://www.ufrgs.br/site/noticias/',
        'https://www.ufrgs.br/noticias/',
    ]
    articles: list[Article] = []
    for url in urls:
        soup = get_soup(url)
        if not soup:
            continue
        articles += extract_titles(soup, [
            'article h2 a', 'article h3 a',
            '.entry-title a', '.field-content a',
            '.views-field-title a', 'h2 a', 'h3 a',
        ], BASE, 'UFRGS Notícias', 'RS')
        if articles:
            break
        time.sleep(DELAY)
    return dedupe(articles)


def scrape_theconversation() -> list[Article]:
    """theconversation.com/br — análise acadêmica em português."""
    BASE = 'https://theconversation.com'
    soup = get_soup('https://theconversation.com/br')
    if not soup:
        return []
    return dedupe(extract_titles(soup, [
        'article h2 a', 'article h3 a',
        '[itemprop="headline"] a',
        '.article-title a', '.story-title a',
        'h2.title a', 'h3.title a',
    ], BASE, 'The Conversation BR', 'Nacional'))


def scrape_estado_rs() -> list[Article]:
    """estado.rs.gov.br — portal de notícias do Governo do RS."""
    BASE = 'https://estado.rs.gov.br'
    urls = [
        'https://estado.rs.gov.br/noticias',
        'https://estado.rs.gov.br/lista-de-noticias',
    ]
    articles: list[Article] = []
    for url in urls:
        soup = get_soup(url)
        if not soup:
            continue
        articles += extract_titles(soup, [
            '.news-list a', '.list-news a',
            'article h2 a', 'article h3 a',
            '.noticia-titulo a', '.field-content a',
            'h2 a[href*="estado.rs.gov.br"]',
            'h3 a[href*="estado.rs.gov.br"]',
            '.views-row a',
        ], BASE, 'Estado RS', 'RS')
        if articles:
            break
        time.sleep(DELAY)
    return dedupe(articles)


def scrape_prefeitura_poa() -> list[Article]:
    """prefeitura.poa.br/noticias — notícias da Prefeitura de Porto Alegre (Drupal)."""
    BASE = 'https://prefeitura.poa.br'
    soup = get_soup('https://prefeitura.poa.br/noticias')
    if not soup:
        return []
    return dedupe(extract_titles(soup, [
        '.views-field-title a',
        '.field-content a',
        'article h2 a', 'article h3 a',
        '.view-content h3 a', '.view-content h2 a',
        'span.field-content a',
        'td a[href*="/noticias/"]',
    ], BASE, 'Prefeitura POA', 'RS'))


def scrape_terra() -> list[Article]:
    """terra.com.br — portal nacional de notícias."""
    BASE = 'https://www.terra.com.br'
    soup = get_soup('https://www.terra.com.br/noticias/')
    if not soup:
        return []
    articles = extract_titles(soup, [
        '.card__title a', '.card-title a',
        'article h2 a', 'article h3 a',
        '[data-type="news"] h2 a', '[data-type="news"] h3 a',
        'h2.title a', 'h3.title a',
    ], BASE, 'Terra', 'Nacional')
    # fallback: links de notícias por padrão de URL
    if not articles:
        for a in soup.find_all('a', href=True):
            link = abs_url(a['href'], BASE)
            if re.match(r'https://www\.terra\.com\.br/.+/\w+-\w+', link):
                title = a.get_text(strip=True)
                if len(title) >= 20:
                    articles.append(Article('Terra', title, link, group='Nacional'))
    return dedupe(articles)


def scrape_diario_gaucho() -> list[Article]:
    """Diário Gaúcho — popular gaúcho, parte do grupo RBS/GZH."""
    BASE = 'https://gauchazh.clicrbs.com.br'
    soup = get_soup('https://gauchazh.clicrbs.com.br/diario-gaucho/')
    if not soup:
        return []
    return dedupe(extract_titles(soup, [
        'h2 a[href*="clicrbs.com.br"]',
        'h3 a[href*="clicrbs.com.br"]',
        'article h2 a', 'article h3 a',
        '[class*="title"] a[href*="clicrbs.com.br"]',
        '[class*="headline"] a',
    ], BASE, 'Diário Gaúcho', 'RS'))


def scrape_google_news(query: str = 'Porto Alegre RS') -> list[Article]:
    """Google News via RSS público (sem API key necessária)."""
    url = (
        f'https://news.google.com/rss/search'
        f'?q={quote(query)}&hl=pt-BR&gl=BR&ceid=BR:pt-419'
    )
    articles: list[Article] = []
    try:
        r = requests.get(url, headers=HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
        root = ET.fromstring(r.content)
        for item in root.findall('.//item'):
            title    = (item.findtext('title') or '').strip()
            link     = (item.findtext('link')  or '').strip()
            pub_date = item.findtext('pubDate') or ''
            desc     = (item.findtext('description') or '').strip()
            # Remove a fonte do título (Google News: "Título - Fonte")
            clean_title = re.sub(r'\s*-\s*[^-]+$', '', title).strip() or title
            date = parse_rfc2822(pub_date)
            if clean_title and link:
                articles.append(Article(
                    source='Google News', title=clean_title,
                    link=link, date=date,
                    description=BeautifulSoup(desc, 'html.parser').get_text()[:200],
                    group='Nacional',
                ))
    except Exception as e:
        log.warning(f"Erro Google News: {e}")
    return articles


def scrape_yahoo_noticias() -> list[Article]:
    """Yahoo Notícias Brasil — portal nacional.
    Nota: renderizado por JS, mas a listagem de categorias é acessível via HTML."""
    BASE = 'https://br.noticias.yahoo.com'
    # Yahoo costuma ter a lista de artigos em tags <a> com classe específica
    soup = get_soup(BASE + '/')
    if not soup:
        return []
    articles: list[Article] = []
    seen: set[str] = set()
    # Padrões de URL de artigos do Yahoo News Brasil
    article_pattern = re.compile(
        r'https?://(br\.noticias\.yahoo\.com|br\.yahoo\.com/news)/'
    )
    for a in soup.find_all('a', href=True):
        link = abs_url(a['href'], BASE)
        if not article_pattern.match(link):
            continue
        title = a.get_text(separator=' ', strip=True)
        if len(title) < 20 or link in seen:
            continue
        seen.add(link)
        articles.append(Article('Yahoo Notícias', title, link, group='Nacional'))
    return dedupe(articles)


# ─────────────────────────────────────────────────────────────────────────────
# Runner principal
# ─────────────────────────────────────────────────────────────────────────────

SCRAPERS: list[tuple[str, callable]] = [
    ('Google News (POA)',    lambda: scrape_google_news('Porto Alegre RS')),
    ('Google News (RS)',     lambda: scrape_google_news('Rio Grande do Sul')),
    ('Nonada',               scrape_nonada),
    ('Sul21',                scrape_sul21),
    ('UFRGS Jornal',         scrape_ufrgs_jornal),
    ('UFRGS Notícias',       scrape_ufrgs_noticias),
    ('The Conversation BR',  scrape_theconversation),
    ('Estado RS',            scrape_estado_rs),
    ('Prefeitura POA',       scrape_prefeitura_poa),
    ('Terra',                scrape_terra),
    ('Diário Gaúcho',        scrape_diario_gaucho),
    ('Yahoo Notícias',       scrape_yahoo_noticias),
]


def run_all(verbose: bool = True) -> list[Article]:
    all_articles: list[Article] = []
    total_sources = len(SCRAPERS)
    for idx, (name, fn) in enumerate(SCRAPERS, 1):
        log.info(f"[{idx}/{total_sources}] Buscando {name}…")
        try:
            arts = fn()
            log.info(f"  → {len(arts)} artigo(s)")
            all_articles.extend(arts)
        except Exception as e:
            log.error(f"  → ERRO em {name}: {e}")
        if idx < total_sources:
            time.sleep(DELAY)
    return all_articles


# ─────────────────────────────────────────────────────────────────────────────
# Exportação para XLSX
# ─────────────────────────────────────────────────────────────────────────────

# Paleta (alinhada ao monitor-matinal.html)
COL_ACCENT  = '42C8DC'
COL_DARK    = '111110'
COL_WHITE   = 'FFFFFF'
COL_LIGHT   = 'F2F1EE'
COL_DIM     = '888880'

SOURCE_COLORS: dict[str, str] = {
    'Nonada':              'D32F2F',
    'Sul21':               '00838F',
    'UFRGS Jornal':        '1565C0',
    'UFRGS Notícias':      '0277BD',
    'The Conversation BR': '6A1B9A',
    'Estado RS':           '2E7D32',
    'Prefeitura POA':      'AD1457',
    'Terra':               'E65100',
    'Diário Gaúcho':       'F57F17',
    'Google News':         '1A73E8',
    'Google News (POA)':   '1A73E8',
    'Google News (RS)':    '1565C0',
    'Yahoo Notícias':      '6200EA',
}


def _header_row(ws, cols: list[str]) -> None:
    for col, label in enumerate(cols, 1):
        c = ws.cell(1, col, label)
        c.font      = Font(bold=True, color=COL_WHITE, size=10)
        c.fill      = PatternFill('solid', fgColor=COL_DARK)
        c.alignment = Alignment(vertical='center', wrap_text=False)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 18


def _hyperlink_cell(ws, row: int, col: int, url: str, text: str = '') -> None:
    c = ws.cell(row, col, text or url)
    if url:
        c.hyperlink = url
        c.style = 'Hyperlink'


def _fmt_date(d: Optional[datetime.datetime]) -> str:
    if not d:
        return '—'
    # Remove timezone para Excel ficar feliz
    if d.tzinfo:
        d = d.astimezone(datetime.timezone.utc).replace(tzinfo=None)
    return d.strftime('%d/%m/%Y %H:%M')


def export_xlsx(articles: list[Article], filepath: str = '') -> str:
    if not filepath:
        ds  = datetime.datetime.now().strftime('%Y-%m-%d')
        filepath = f'matinal-scraping-{ds}.xlsx'

    wb  = openpyxl.Workbook()
    wb.remove(wb.active)
    now = datetime.datetime.now()

    sorted_arts = sorted(
        articles,
        key=lambda a: a.date or datetime.datetime.min,
        reverse=True,
    )

    # ── Aba: Resumo ───────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Resumo')
    ws_sum.column_dimensions['A'].width = 26
    ws_sum.column_dimensions['B'].width = 14
    ws_sum.column_dimensions['C'].width = 8

    ws_sum.cell(1, 1, 'Matinal · Monitor de Notícias').font = Font(
        bold=True, size=14, color=COL_ACCENT)
    ws_sum.cell(2, 1, 'Scraping de fontes sem RSS público').font = Font(
        italic=True, color=COL_DIM)
    ws_sum.cell(3, 1, f'Gerado em: {now.strftime("%d/%m/%Y %H:%M")}').font = Font(
        color=COL_DIM)
    ws_sum.cell(4, 1, f'Total de artigos coletados: {len(articles)}').font = Font(bold=True)

    _header_row(ws_sum, ['Fonte', 'Grupo', 'Artigos'])

    counts: dict[str, dict] = {}
    for a in articles:
        if a.source not in counts:
            counts[a.source] = {'group': a.group, 'n': 0}
        counts[a.source]['n'] += 1

    for row, (src, info) in enumerate(sorted(counts.items()), 2):
        color = SOURCE_COLORS.get(src, COL_ACCENT)
        c = ws_sum.cell(row, 1, src)
        c.fill = PatternFill('solid', fgColor=color + '22')
        ws_sum.cell(row, 2, info['group'])
        ws_sum.cell(row, 3, info['n']).alignment = Alignment(horizontal='center')

    # ── Aba: Todas as notícias ────────────────────────────────────────────────
    ws_all = wb.create_sheet('Todas as Notícias')
    ws_all.column_dimensions['A'].width = 24
    ws_all.column_dimensions['B'].width = 12
    ws_all.column_dimensions['C'].width = 82
    ws_all.column_dimensions['D'].width = 18

    _header_row(ws_all, ['Fonte', 'Grupo', 'Manchete', 'Data/Hora'])

    for row, art in enumerate(sorted_arts, 2):
        color = SOURCE_COLORS.get(art.source, COL_ACCENT)
        c = ws_all.cell(row, 1, art.source)
        c.fill = PatternFill('solid', fgColor=color + '22')
        ws_all.cell(row, 2, art.group)
        # Manchete com hyperlink
        title_cell = ws_all.cell(row, 3, art.title)
        if art.link:
            title_cell.hyperlink = art.link
            title_cell.style = 'Hyperlink'
        ws_all.cell(row, 4, _fmt_date(art.date))

    # ── Uma aba por fonte ─────────────────────────────────────────────────────
    sources = sorted(set(a.source for a in sorted_arts))
    for src in sources:
        safe  = re.sub(r'[\\/*?:\[\]]', '', src)[:31]
        ws    = wb.create_sheet(safe)
        color = SOURCE_COLORS.get(src, COL_ACCENT)

        ws.column_dimensions['A'].width = 85
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 18

        # Cabeçalho colorido por fonte
        for col, label in enumerate(['Manchete', 'Link', 'Data/Hora'], 1):
            c = ws.cell(1, col, label)
            c.font      = Font(bold=True, color=COL_WHITE, size=10)
            c.fill      = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(vertical='center')
        ws.freeze_panes = 'A2'
        ws.row_dimensions[1].height = 18

        src_arts = [a for a in sorted_arts if a.source == src]
        for row, art in enumerate(src_arts, 2):
            title_cell = ws.cell(row, 1, art.title)
            if art.link:
                title_cell.hyperlink = art.link
                title_cell.style = 'Hyperlink'
            _hyperlink_cell(ws, row, 2, art.link)
            ws.cell(row, 3, _fmt_date(art.date))

    wb.save(filepath)
    log.info(f"Planilha salva: {filepath}")
    return filepath


# ─────────────────────────────────────────────────────────────────────────────
# Ponto de entrada
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    print("=" * 55)
    print("  Matinal · Monitor de Notícias — Scraper")
    print("=" * 55)
    articles = run_all()
    print(f"\n{'─'*55}")
    print(f"Total coletado: {len(articles)} artigos")
    print(f"Fontes: {len(set(a.source for a in articles))}")
    if articles:
        path = export_xlsx(articles)
        print(f"Planilha gerada: {path}")
    else:
        print("Nenhum artigo coletado.")
    print("=" * 55)


if __name__ == '__main__':
    main()
